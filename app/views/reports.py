import io

import openpyxl
from flask import Blueprint, render_template, request, flash, redirect, url_for, send_file
from flask_login import login_required
from app.models import Project, Invoice, SubcontractorPayment, Client, Subcontractor
from app import db
from sqlalchemy import func

reports_bp = Blueprint('reports', __name__, url_prefix='/reports')


@reports_bp.route('/')
@login_required
def index():
    # Per-client összesítő
    clients = Client.query.filter_by(is_active=True).all()
    client_stats = []
    for c in clients:
        invoiced = sum(float(i.amount) for p in c.projects for i in p.invoices if i.direction == 'outgoing')
        received = sum(float(i.amount) for p in c.projects for i in p.invoices if i.direction == 'outgoing' and i.status == 'paid')
        client_stats.append({
            'client': c,
            'invoiced': invoiced,
            'received': received,
            'outstanding': invoiced - received,
            'project_count': len(c.projects)
        })

    # Per-subcontractor összesítő
    subs = Subcontractor.query.filter_by(is_active=True).all()
    sub_stats = []
    for s in subs:
        total = sum(float(p.amount) for p in s.payments)
        paid = sum(float(p.amount) for p in s.payments if p.status == 'paid')
        sub_stats.append({
            'sub': s,
            'total': total,
            'paid': paid,
            'pending': total - paid
        })

    return render_template('reports/index.html',
                           client_stats=client_stats,
                           sub_stats=sub_stats)


@reports_bp.route('/project/<int:id>')
@login_required
def project_report(id):
    project = Project.query.get_or_404(id)
    return render_template('reports/project.html', project=project)


@reports_bp.route('/export')
@login_required
def export():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Projektek'
    ws.append(['Projekt', 'Ügyfél', 'Státusz', 'Szerződéses érték', 'Kiállított', 'Befolyt', 'Kintlévőség', 'Kezdés', 'Befejezés'])
    for p in Project.query.order_by(Project.name).all():
        ws.append([
            p.name,
            p.client.name if p.client else '',
            p.status,
            float(p.contract_value or 0),
            p.total_invoiced,
            p.total_received,
            p.outstanding,
            p.start_date,
            p.end_date,
        ])

    ws = wb.create_sheet('Számlák')
    ws.append(['Számlaszám', 'Projekt', 'Ügyfél', 'Irány', 'Nettó', 'ÁFA %', 'Bruttó', 'Kiállítás', 'Esedékesség', 'Fizetve', 'Státusz'])
    for i in Invoice.query.order_by(Invoice.created_at.desc()).all():
        ws.append([
            i.invoice_number,
            i.project.name if i.project else '',
            i.client.name if i.client else '',
            i.direction,
            float(i.amount or 0),
            float(i.vat_rate or 0),
            float(i.amount_with_vat or i.amount or 0),
            i.issue_date,
            i.due_date,
            i.paid_date,
            i.status,
        ])

    ws = wb.create_sheet('Alvállalkozói kifizetések')
    ws.append(['Projekt', 'Alvállalkozó', 'Összeg', 'Leírás', 'Számla ref.', 'Kiállítás', 'Esedékesség', 'Fizetve', 'Státusz'])
    for p in SubcontractorPayment.query.order_by(SubcontractorPayment.created_at.desc()).all():
        ws.append([
            p.project.name if p.project else '',
            p.subcontractor.name if p.subcontractor else '',
            float(p.amount or 0),
            p.description or '',
            p.invoice_ref or '',
            p.issue_date,
            p.due_date,
            p.paid_date,
            p.status,
        ])

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(vertical='top')
        for column_cells in sheet.columns:
            length = max(len(str(cell.value or '')) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 42)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name='solvior-riportok.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
